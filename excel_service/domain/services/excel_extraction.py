from openpyxl import load_workbook
from fastapi import UploadFile
from io import BytesIO
from pymongo.errors import DuplicateKeyError
from config.db import get_clients
import logging

logger = logging.getLogger(__name__)


def process_excel(file: UploadFile, col_name, col_number, col_review, col_category,
                  col_adress, col_web, col_mail, col_latitude, col_length, col_ubication):
    content = file.file.read()
    wb = load_workbook(filename=BytesIO(content))
    sheet = wb.active
    clients = []

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    col_index = {name: idx for idx, name in enumerate(headers)}
    required_cols = [col_name, col_number, col_review, col_category, col_adress,
                     col_web, col_mail, col_latitude, col_length, col_ubication]

    for col in required_cols:
        if col not in col_index:
            raise ValueError(f"Columna '{col}' no encontrada en el Excel.")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        clients.append({
            "name": row[col_index[col_name]],
            "phone": row[col_index[col_number]],
            "review": row[col_index[col_review]],
            "category": row[col_index[col_category]],
            "adress": row[col_index[col_adress]],
            "web": row[col_index[col_web]],
            "mail": row[col_index[col_mail]],
            "latitude": row[col_index[col_latitude]],
            "length": row[col_index[col_length]],
            "ubication": row[col_index[col_ubication]],
            "state": "pending"
        })
    return clients


def normalize(value):
    return str(value).strip().lower() if value else ""

def build_empty_key(client):
    return f"{normalize(client.get('name'))}-{normalize(client.get('adress'))}-{normalize(client.get('ubication'))}"

async def save_clients(client_list: list[dict]):
    try:
        clients_collection = get_clients()
        if not client_list:
            return

        seen_phones = set()
        seen_empty_keys = set()
        filtered_clients = []

        for c in client_list:
            phone = normalize(c.get("phone"))

            if phone == "":
                key = build_empty_key(c)
                if key not in seen_empty_keys:
                    seen_empty_keys.add(key)
                    filtered_clients.append(c)
            elif phone not in seen_phones:
                seen_phones.add(phone)
                filtered_clients.append(c)
            else:
                # teléfono duplicado en el mismo archivo
                logger.warning(f"Teléfono duplicado en Excel: {phone} - Cliente: {c.get('name')}")

        # Verificar duplicados en la base de datos
        phones_to_check = [normalize(c["phone"]) for c in filtered_clients if normalize(c.get("phone")) != ""]
        existing_phones = {
            client["phone"]
            for client in clients_collection.find(
                {"phone": {"$in": phones_to_check}},
                {"phone": 1}
            )
        }

        new_clients = []
        for c in filtered_clients:
            phone = normalize(c.get("phone"))
            if phone == "":
                new_clients.append(c)
            elif phone not in existing_phones:
                new_clients.append(c)
            else:
                logger.warning(f"Teléfono ya existe en base de datos: {phone} - Cliente: {c.get('name')}")

        if new_clients:
            result = clients_collection.insert_many(new_clients)
            logger.info(f"{len(result.inserted_ids)} clientes insertados correctamente.")
        else:
            logger.info("No hay clientes nuevos para insertar.")

    except Exception as e:
        logger.error(f'Error al guardar clientes: {e}')
        raise
